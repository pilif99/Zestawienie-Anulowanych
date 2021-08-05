
import openpyxl
import pyautogui as gui, time
import xlsxwriter
import pandas
import eden
import numpy as np
import data


class Pobieranie_Danych:

    tryb_testowy = False

    def __init__(self):

        super().__init__()

        if not Pobieranie_Danych.tryb_testowy:

            writer = pandas.ExcelWriter('zestawienie_anulowanych_' + str(data.Data()) + '.xlsx', engine='xlsxwriter', options={'strings_to_formulas': False})

            # stwórz zrealizowane i anulowane
        
            df1 = eden.Eden('anulowane')
            df2 = eden.Eden('zrealizowane')

            # usuń niepotrzebne marki

            for i in ('BROGER', 'OZONE', 'REBELHORN'):

                df1.drop(df1[df1.iProducent == i].index, inplace = True)
                df2.drop(df2[df2.Producent == i].index, inplace = True)

            # wrzucenie zrealizowanych i anulowanych do excela

            df1.to_excel(writer, sheet_name='Anulowane', index=False)
            df2.to_excel(writer, sheet_name='Zrealizowane', index=False)

            # stworzenie zrealizowanych2 i anulowanych2

            df3 = df1[['dIndeks', 'fNazwa', 'gNiezrealizowanaIlosc', 'gNiezrealizowanaWartosc', 'kStatusP', 'iProducent']]
            df4 = df2[['Grupa', 'Opis', 'Ilosc', 'Sprzedaż', 'Status', 'Producent']]

            # stworzenie indeksu
        
            kolumny = ['Indeks', 'Nazwa', 'Ilość', 'Wartość', 'Status', 'Producent']
            kolumny2 = ['Indeks', 'Nazwa', 'Status', 'Producent', 'Ilość Zrealizowana', 'Wartość Zrealizowana', 'Ilość Anulowana', 'Wartość Anulowana']

            df3.columns = kolumny
            df4.columns = kolumny

            df_indeks = pandas.concat([df3, df4])
            df_indeks = df_indeks.drop(['Ilość', 'Wartość'], axis = 'columns')
            df_indeks = df_indeks.drop_duplicates(subset=['Indeks'])
            df_indeks = df_indeks.reset_index(drop = True)

            # uzupełnianie indeksu

            df_indeks = pandas.merge(df_indeks, df4, how = 'left', on = ['Indeks'])
            df_indeks.replace(np.nan, 0, regex=True, inplace=True)
            df_indeks = df_indeks.drop(['Nazwa_y', 'Status_y', 'Producent_y'], axis = 'columns')
            df_indeks = pandas.merge(df_indeks, df3.groupby('Indeks')['Ilość'].sum(), how = 'left', on = ['Indeks'])
            df_indeks = pandas.merge(df_indeks, df3.groupby('Indeks')['Wartość'].sum(), how = 'left', on = ['Indeks'])
            df_indeks = df_indeks.replace(np.nan, 0, regex=True)
            df_indeks.columns = kolumny2

            # tworzenie kolumny modelokoloru i wrzucenie do excela

            df_indeks['Modelokolor'] = df_indeks['Indeks'].apply(lambda row: ''.join(list(row)[:[i for i, n in enumerate(list(row)) if n == '_'][1]]) if list(row).count('_') == 2 else row)
            df_indeks = df_indeks[['Indeks', 'Modelokolor', 'Nazwa', 'Status', 'Producent', 'Ilość Zrealizowana', 'Wartość Zrealizowana', 'Ilość Anulowana', 'Wartość Anulowana']]
            df_indeks.to_excel(writer, sheet_name='Indeks', index=False)
            # tworzenie modelokoloru

            df_modelokolor = df_indeks[['Modelokolor', 'Nazwa', 'Status', 'Producent']]
            df_modelokolor = df_modelokolor.drop_duplicates(subset=['Modelokolor'])
            df_modelokolor = pandas.merge(df_modelokolor, df_indeks.groupby('Modelokolor')['Ilość Zrealizowana'].sum(), how = 'left', on = ['Modelokolor'])
            df_modelokolor = pandas.merge(df_modelokolor, df_indeks.groupby('Modelokolor')['Wartość Zrealizowana'].sum(), how = 'left', on = ['Modelokolor'])
            df_modelokolor = pandas.merge(df_modelokolor, df_indeks.groupby('Modelokolor')['Ilość Anulowana'].sum(), how = 'left', on = ['Modelokolor'])
            df_modelokolor = pandas.merge(df_modelokolor, df_indeks.groupby('Modelokolor')['Wartość Anulowana'].sum(), how = 'left', on = ['Modelokolor'])
            df_modelokolor.to_excel(writer, sheet_name='Modelokolor', index=False)

            # tworzenie marki

            df_marka = df_modelokolor[['Producent']]
            df_marka = df_marka.drop_duplicates(subset = ['Producent'])
            df_marka = pandas.merge(df_marka, df_modelokolor.groupby('Producent')['Ilość Zrealizowana'].sum(), how = 'left', on = ['Producent'])
            df_marka = pandas.merge(df_marka, df_modelokolor.groupby('Producent')['Wartość Zrealizowana'].sum(), how = 'left', on = ['Producent'])
            df_marka = pandas.merge(df_marka, df_modelokolor.groupby('Producent')['Ilość Anulowana'].sum(), how = 'left', on = ['Producent'])
            df_marka = pandas.merge(df_marka, df_modelokolor.groupby('Producent')['Wartość Anulowana'].sum(), how = 'left', on = ['Producent'])
            df_marka.to_excel(writer, sheet_name='Marka', index=False)
        
        else:

            # pobranie danych z excela

            df1 = pandas.read_excel('zestawienie_anulowanych_' + str(data.Data()) + '.xlsx', sheet_name = 'Anulowane')
            df2 = pandas.read_excel('zestawienie_anulowanych_' + str(data.Data()) + '.xlsx', sheet_name = 'Zrealizowane')
            df_indeks = pandas.read_excel('zestawienie_anulowanych_' + str(data.Data()) + '.xlsx', sheet_name = 'Indeks')
            df_modelokolor = pandas.read_excel('zestawienie_anulowanych_' + str(data.Data()) + '.xlsx', sheet_name = 'Modelokolor')
            df_marka = pandas.read_excel('zestawienie_anulowanych_' + str(data.Data()) + '.xlsx', sheet_name = 'Marka')

            writer = pandas.ExcelWriter('zestawienie_anulowanych_' + str(data.Data()) + '.xlsx', engine='xlsxwriter', options={'strings_to_formulas': False})
            df1.to_excel(writer, sheet_name='Anulowane', index=False)
            df2.to_excel(writer, sheet_name='Zrealizowane', index=False)
            df_indeks.to_excel(writer, sheet_name='Indeks', index=False)
            df_modelokolor.to_excel(writer, sheet_name='Modelokolor', index=False)
            df_marka.to_excel(writer, sheet_name='Marka', index=False)

        workbook = writer.book

        for i, j in {'Anulowane': df1, 'Zrealizowane': df2, 'Indeks': df_indeks, 'Modelokolor': df_modelokolor, 'Marka': df_marka}.items():

            sheet = writer.sheets[i]
            print(chr(len(j.columns)+64))
            formatShadeRows = workbook.add_format({'bg_color': '#e9e9e9',
                                                    'font_color': 'black'})
            sheet.conditional_format(f'A1:{chr(len(j.columns)+64)}{len(j.index)+1}',{'type': 'formula', 
                                                                                        'criteria': '=MOD(ROW(),2) = 0', 
                                                                                        'format': formatShadeRows})

        writer.save()

        wb = openpyxl.load_workbook('zestawienie_anulowanych_' + str(data.Data()) + '.xlsx')

        for ws in wb.worksheets:
            for i in [cell.column for cell in ws[1]]:
                for cell in ws[openpyxl.utils.cell.get_column_letter(i)]:
                    cell.number_format = openpyxl.styles.numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        for ws in wb.worksheets[-3:]:
    
            for i in [cell.column for cell in ws[1]][-5:]:
        
                ws.column_dimensions[openpyxl.utils.cell.get_column_letter(i)].width = len(ws.cell(1, i).value) + 2

        ws = wb['Marka']

        literki = list('BCDE')

        ostatni = ws.max_row

        for i in range(4):

            ws.cell(ostatni + 2, i + 2).value = sum([ws.cell(x + 2, i + 2).value for x in range(ostatni - 1)])
            ws.cell(ostatni + 2, i + 2).number_format = openpyxl.styles.numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        wb.save('zestawienie_anulowanych_' + str(data.Data()) + '.xlsx')